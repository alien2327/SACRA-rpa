import os
import sys
sys.path.append(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))

import openpyxl
import pandas as pd
from pandas.io import sql

from sqlalchemy import create_engine
import mysql.connector

import calendar

def read_dir(pathname):
    engine = create_engine("mysql://root:Ryou017273@localhost/test", encoding='utf-8', echo=False)
    path = "./" + pathname + "/"
    worker = {}
    files = os.listdir(path)
    for worker_name in files:
        years = os.listdir(path+worker_name)
        for sheet in years:
            print(f"Processing with {worker_name} {sheet}")
            excel = os.listdir(path+worker_name+"/"+sheet)
            work_data = read_excel(sheet, path+worker_name+"/"+sheet, excel)
            #work_data.to_sql(worker_name, engine, index=False, schema='test', if_exists='append')
            worker.update({worker_name:work_data})
    return pd.concat(worker)

def read_excel(year, path, pathlist):
    global minc, maxc, minr, maxr
    month_index = [i+1 for i in range(12)]
    month_data = []
    for file in pathlist:
        workdata = []
        wb = openpyxl.load_workbook(filename = path + '/' + file, data_only=True)
        ws = wb[wb.sheetnames[0]]
        minc = ws.min_column
        maxc = ws.max_column
        minr = ws.min_row
        maxr = ws.max_row
        columns = ['作業者印', '業務内容', '定められた勤務時間', '減額時間', '(休憩)/超過(1)', '(休憩)/超過(2)', '超過(100/100)', '超過(125/100)', '超過(25/100)', 'その他', '監督・命令者認印', '従事者印']
        index = read_day(ws, minc) + read_day(ws, minc+14)

        data = ws.iter_rows(min_row=minr+3, min_col=minc, max_col=minc+12)
        values = []
        for row in data:
            if row[0].value:
                for col in row[1:]:
                    values.append(col.value)
            elif row[0].value is None:
                values[4] = [values[4], row[5].value]
                values[5] = [values[5], row[6].value]
                workdata.append(values)
                values = []

        data = ws.iter_rows(min_row=minr+3, min_col=minc+14, max_col=maxc)
        values = []
        for row in data:
            try:
                if row[0].value and row[0].value != "計":
                    for col in row[1:]:
                        values.append(col.value)
                elif row[0].value is None:
                    values[4] = [values[4], row[5].value]
                    values[5] = [values[5], row[6].value]
                    workdata.append(values)
                    values = []
            except IndexError:
                continue
        
        df_work = pd.DataFrame(workdata, index=index, columns=columns)

        data = list(range(11))
        result = map(lambda x: 0 if x == 0 or x == 3 or x == 4 or x == 9 or x == 10 else df_work.sum()[x], data)
        result = list(result)
        result.insert(0, 0)
        df_result = pd.DataFrame([result], index=["計"], columns=columns)
        df_work = df_work.append(df_result)
        month_data.append(df_work)
    month_dict = dict(zip(month_index, month_data))
    month_df = pd.concat(month_dict)
    return month_df

def read_day(ws, start_point):
    global minr, maxr
    index = []
    for i in range(minr+3, maxr):
        val = ws.cell(i, start_point).value
        if val == None or val == "計":
            continue
        else:
            index.append(val)
    return index

def link_database(dataframe=None):
    connect = mysql.connector.connect(user='root', password='Ryou017273', host='localhost', database='', charset='utf8')
    print(connect.is_connected())

if __name__ == '__main__':
    worker = read_dir(sys.argv[1])
    print(worker)
    #link_database()