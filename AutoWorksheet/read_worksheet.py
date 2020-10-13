import os
import sys
sys.path.append(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))

import numpy as np
import openpyxl
import pandas as pd

import calendar

# 以下のディレクトリは探索しない
excludes=['test', '__pycache__']

def read_dir(name):
    path_dir = f'./{name}/'
    try:
        year = sys.argv[2]
        print('read files in {}{}/' .format(path_dir, year))
        read_worksheet(name, int(year))
    except:
        files_name = os.listdir(path_dir)
        years = [f for f in files_name if os.path.isdir(os.path.join(path_dir,f))]
        for year in years:
            print('read files in {}{}/' .format(path_dir, year))
            read_worksheet(name, int(year))

def read_worksheet(name, year):
    data = []
    for m in range(12):
        month = m+1
        date_max = calendar.monthrange(year, month)[1]

        path_xlsx = f'./{name}/{year}/{month}_work.xlsx'
        wb = openpyxl.load_workbook(path_xlsx, data_only=True)
        ws = wb[wb.sheetnames[0]]

        c_1 = ['B','C','D','E','F','G','H','I','J','K','L','M','N']
        c_2 = ['P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB']
        col = [c_1, c_2]
        early = 16
        late = 15
        r_1 = [5+2*i for i in range(early)]
        r_2 = [5+2*i for i in range(late)]
        row = [r_1, r_2]

        for tag in range(2):
            for i,r in enumerate(row[tag]):
                data.append([month])
                for j,c in enumerate(col[tag]):
                    if j == 1 or j == 11 or j == 12:
                        continue
                    elif j == 5 or j == 6:
                        data[-1].append(ws[c+str(r)].value)
                        data[-1].append(ws[c+str(r+1)].value)
                    else:
                        data[-1].append(ws[c+str(r)].value)
                if i+early*tag+1 == date_max:
                    break
        tag = 1
        r = 35
        data.append([month])
        for j,c in enumerate(col[tag]):
            if j == 1 or j == 11 or j == 12:
                continue
            elif j == 0:
                data[-1].append('sum')
            elif j == 5 or j == 6:
                data[-1].append(ws[c+str(r)].value)
                data[-1].append(ws[c+str(r+1)].value)
            else:
                data[-1].append(ws[c+str(r)].value)
    df = pd.DataFrame(data)
    df.to_csv(f'./{name}/{year}_workdata.csv')

if __name__ == '__main__':
    path = './'
    try:
        name = sys.argv[1]
        print('read files in {}{}/' .format(path, name))
        read_dir(name)
    except:
        files = os.listdir(path)
        names = []
        for f in files:
            if os.path.isdir(os.path.join(path,f)) and not f in excludes:
                names.append(f)
        for name in names:
            print('read files in {}{}/' .format(path,name))
            read_dir(name)
