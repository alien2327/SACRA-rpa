import os
import sys
sys.path.append(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))

import numpy as np
import openpyxl
import pandas as pd

import calendar

def read_dir(path,name):
    path_dir = f'{path}{name}/'
    try:
        year = sys.argv[2]
        print('read files in {}{}/' .format(path_dir, year))
        read_worksheet(path_dir, int(year))
    except:
        files_name = os.listdir(path_dir)
        years = [f for f in files_name if os.path.isdir(os.path.join(path_dir,f))]
        for year in years:
            print('read files in {}{}/' .format(path_dir, year))
            read_worksheet(path_dir, int(year))

def read_worksheet(path_dir, year):
    # 直下は共通なので def read_worksheet の外に置いてもよいか
    c_1 = ['B','C','D','E','F','G','H','I','J','K','L','M','N']
    c_2 = ['P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB']
    col = [c_1, c_2]
    early = 16
    late = 15
    r_1 = [5+2*i for i in range(early)]
    r_2 = [5+2*i for i in range(late)]
    row = [r_1, r_2]
    skip = [1, 11, 12]
    time = [3, 4, 7, 8, 9 ,10]
    dual = [5, 6]

    data = []
    for m in range(12):
        month = m+1
        date_max = calendar.monthrange(year, month)[1]

        path_xlsx = f'{path_dir}/{year}/{month}_work.xlsx'
        wb = openpyxl.load_workbook(path_xlsx, data_only=True)
        ws = wb[wb.sheetnames[0]]

        sums = [0 for i in range(13)]
        for tag in range(2):
            for i,r in enumerate(row[tag]):
                value = [month]
                for j,c in enumerate(col[tag]):
                    if j in skip:
                        continue
                    elif j in time:
                        try:
                            value.append(float(ws[c+str(r)].value))
                        except:
                            value.append(0.0)
                        else:
                            sums[j] += value[-1]
                    elif j in dual:
                        value.append(ws[c+str(r)].value)
                        value.append(ws[c+str(r+1)].value)
                    else:
                        value.append(ws[c+str(r)].value)

                data.append(value)
                if i+early*tag+1 == date_max:
                    break

        # 集計行
        tag = 1
        r = 35
        data.append([month])
        for j,c in enumerate(col[tag]):
            if j in skip:
                continue
            elif j in time:
                data[-1].append(sums[j])
            elif j in dual:
                data[-1].append(ws[c+str(r)].value)
                data[-1].append(ws[c+str(r+1)].value)
            else:
                data[-1].append(ws[c+str(r)].value)

    columns = ['月', '日', '業内', '定められた勤務時間', '減額時間', '休憩時間(1)', '超過勤務時間(1)', '休憩時間(2)', '超過勤務時間(2)', '超過勤務等(100/100)', '超過勤務等(125/100)', '超過勤務等(25/100)', 'その他']
    df = pd.DataFrame(data, columns=columns)
    df.to_csv(f'{path_dir}/{year}_workdata.csv')

if __name__ == '__main__':
    path = './data/'
    try:
        name = sys.argv[1]
        print('read files in {}{}/' .format(path, name))
        read_dir(path,name)
    except:
        files = os.listdir(path)
        names = []
        for f in files:
            if os.path.isdir(os.path.join(path,f)):
                names.append(f)
        for name in names:
            print('read files in {}{}/' .format(path,name))
            read_dir(path,name)
