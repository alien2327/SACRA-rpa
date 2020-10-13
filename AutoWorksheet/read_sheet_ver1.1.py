import os
import sys
sys.path.append(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))

import openpyxl
import pandas as pd

import calendar

def read_dir(pathname):
    path = "./" + pathname + "/"
    worker = {}
    files = os.listdir(path)
    for folder in files:
        years = os.listdir(path+folder)
        worksheet = []
        for sheet in years:
            excel = os.listdir(path+folder+"/"+sheet)
            read_excel(sheet, path+folder+"/"+sheet, excel)
            worksheet.append(excel)
        worker.update({folder:worksheet})
    return worker

def read_excel(year, path, pathlist):
    global minc, maxc, minr, maxr
    for file in pathlist:
        workdata = []
        wb = openpyxl.load_workbook(filename = path + '/' + file, data_only=True)
        ws = wb[wb.sheetnames[0]]
        minc = ws.min_column
        maxc = ws.max_column
        minr = ws.min_row
        maxr = ws.max_row
        columns = ['作業者印', '業務内容', '定められた勤務時間', '減額時間', '(休憩)/超過(1)', '(休憩)/超過(2)', '超過(100/100)', '超過(125/100)', '超過(25/100)', 'その他', '監督・命令者認印', '従事者印']
        index = read_day(ws, minc) + read_day(ws, minc+14)

        first = ws.iter_rows(min_row=minr+3, min_col=minc, max_col=minc+12)
        second = ws.iter_rows(min_row=minr+3, min_col=minc+14, max_col=maxc)

        for row in first:
            values = []
            for col in row[1:]:
                values.append(col.value)
            if row[0].value is None:
                values[4] = [values[4], row[5].value]
                values[5] = [values[5], row[6].value]
                workdata.append(values)
                
            
        for row in second:
            values = []
            if row[0].value is None:
                continue
            for col in row[1:]:
                values.append(col.value)
            workdata.append(values)

        df_work = pd.DataFrame(workdata, index=index, columns=columns)
        print(df_work)

def read_day(ws, start_point):
    global minr, maxr
    index = []
    for i in range(minr+3, maxr):
        val = ws.cell(i, start_point).value
        if val == None:
            continue
        else:
            index.append(val)
    return index

if __name__ == '__main__':
    worker = read_dir(sys.argv[1])