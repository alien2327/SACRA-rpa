import calendar
import datetime
import jpholiday
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side

def isBizDay(year, month, day):
    Date = datetime.date(year, month, day)
    if jpholiday.is_holiday(Date):
        return 0
    else:
        return 1

year = int(input("Insert Year you want to generate. : "))
month = int(input("Insert Month you want to generate. : "))

date_max = calendar.monthrange(year, month)[1]

c_1 = ['B','C','D','E','F','G','H','I','J','K','L','M','N']
c_2 = ['P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB']
col = [c_1, c_2]
early = 16
late = 15
r_1 = [5+2*i for i in range(early)]
r_2 = [5+2*i for i in range(late)]
row = [r_1, r_2]
wb = Workbook()
ws = wb.active

for i in range(36):
    ws.row_dimensions[i+1].height = 18.75

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
top_border = Border(top=Side(style='thin'))
bottom_border = Border(bottom=Side(style='thin'))
slash_border = Border(diagonal=Side(style='thin'),left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
slash_border.diagonalDown=True

cell_alignment = Alignment(horizontal = 'center', vertical ='center', wrap_text=True)

for tag in range(2):
    for c in col[tag]:
        for i in range(2,37):
            cell = ws[c + str(i)]
            cell.alignment = cell_alignment
            cell.border = thin_border

ws['O2'].border = top_border
ws['O36'].border = bottom_border

# Merging Cell
for tag in range(2):
    ws[col[tag][0]+'2'] = '日付'
    ws[col[tag][0]+'2'].font = Font(size = 10)
    ws.merge_cells(col[tag][0]+'2:'+col[tag][0]+'4') #日付
    ws[col[tag][1]+'2'] = '作業者印\n(始業)'
    ws[col[tag][1]+'2'].font = Font(size = 9)
    ws.merge_cells(col[tag][1]+'2:'+col[tag][1]+'4') #作業者印
    ws[col[tag][2]+'2'] = '勤　　　　　務　　　　　内　　　　　容'
    #ws[col[tag][2]+'2'].font = Font(size = 11)
    ws.merge_cells(col[tag][2]+'2:'+col[tag][10]+'2') #勤務内容
    ws[col[tag][2]+'3'] = '業　内'
    #ws[col[tag][2]+'3'].font = Font(size = 11)
    ws.merge_cells(col[tag][2]+'3:'+col[tag][2]+'4') #業内
    ws[col[tag][3]+'3'] = '定められた\n勤務時間'
    ws[col[tag][3]+'3'].font = Font(size = 8)
    ws.merge_cells(col[tag][3]+'3:'+col[tag][3]+'4') #定められた勤務時間
    ws[col[tag][4]+'3'] = '減額時間'
    ws[col[tag][4]+'3'].font = Font(size = 8)
    ws.merge_cells(col[tag][4]+'3:'+col[tag][4]+'4') #減額時間
    ws[col[tag][5]+'3'] = '(休憩時間)'
    ws[col[tag][5]+'3'].font = Font(size = 10)
    ws.merge_cells(col[tag][5]+'3:'+col[tag][6]+'3') #休憩時間
    ws[col[tag][5]+'4'] = '超過勤務時間'
    ws[col[tag][5]+'4'].font = Font(size = 10)
    ws.merge_cells(col[tag][5]+'4:'+col[tag][6]+'4') #超過勤務時間
    ws[col[tag][7]+'3'] = '超過勤務等'
    #ws[col[tag][7]+'3'].font = Font(size = 11)
    ws[col[tag][7]+'4'] = '100/100'
    ws[col[tag][7]+'4'].font = Font(size = 8)
    ws[col[tag][8]+'4'] = '125/100'
    ws[col[tag][8]+'4'].font = Font(size = 8)
    ws[col[tag][9]+'4'] = '25/100'
    ws[col[tag][9]+'4'].font = Font(size = 8)
    ws.merge_cells(col[tag][7]+'3:'+col[tag][9]+'3') #超過勤務等
    ws[col[tag][10]+'3'] = 'その他'
    ws[col[tag][10]+'3'].font = Font(size = 9)
    ws.merge_cells(col[tag][10]+'3:'+col[tag][10]+'4') #その他
    ws[col[tag][11]+'2'] = '監督・命令者\n認印'
    ws[col[tag][11]+'2'].font = Font(size = 9)
    ws.merge_cells(col[tag][11]+'2:'+col[tag][11]+'4') #監督・命令者認印
    ws[col[tag][12]+'2'] = '従事者印\n(超勤)'
    ws[col[tag][12]+'2'].font = Font(size = 9)
    ws.merge_cells(col[tag][12]+'2:'+col[tag][12]+'4') #従事者印

    for j,r in enumerate(row[tag]):
        date = j + 1 + early*tag
        if date > date_max:
            for c in col[tag]:
                range = c + str(r) + ':' + c + str(r+1)
                ws.merge_cells(range)
                ws[c + str(r)].border = slash_border
        else:
            for i,c in enumerate(col[tag]):
                if i == 5 or i == 6:
                    ws[c + str(r)] = '(　：　)'
                    ws[c + str(r+1)] = '：'
                else:
                    if i == 0:
                        ws[c + str(r)] = str(date)
                    if i == 2:
                        if isBizDay(year, month, date) == 0:
                            ws[c + str(r)] = '(祝 日)'
                        elif calendar.weekday(year, month, date) == 5:
                            ws[c + str(r)] = '(土 曜)'
                        elif calendar.weekday(year, month, date) == 6:
                            ws[c + str(r)] = '(日 曜)'
                    range = c + str(r) + ':' + c + str(r+1)
                    ws.merge_cells(range)

#最後の列
tag = 1
for i,c in enumerate(col[tag]):
    if i == 0:
        ws[c + '35'] = '計'
        ws.merge_cells(c + '35:' + col[tag][i+2] + '36')
    if i in [3,4,7,8,9,10]:
        sum = '=SUM(' + col[0][i] + '5:' + col[0][i] + '36,'
        sum += col[1][i] + '5:' + col[1][i] + '34)'
        ws[col[tag][i] + '35'] = sum
        ws[col[tag][i] + '35'].font = Font(size = 14)
        ws.merge_cells(c + '35:' + c + '36')
    if i in [5,11]:
        ws.merge_cells(c + '35:' + col[tag][i+1] + '36')
        ws[c + '35'].border = slash_border

wb.save("worksheet_"+str(year)+"_"+str(month)+".xlsx")
