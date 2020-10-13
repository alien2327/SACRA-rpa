import calendar
import datetime
import jpholiday
from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side

def isBizDay(year, month, day):
    Date = datetime.date(year, month, day)
    if jpholiday.is_holiday(Date):
        return 0
    else:
        return 1

year = input("Insert Year you want to generate. : ")
month = input("Insert Month you want to generate. : ")

c_1 = ['B','C','D','E','F','G','H','I','J','K','L','M','N']
c_2 = ['P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB']
column = [c_1, c_2]

wb = Workbook()
ws = wb.active

for i in range(36):
    ws.row_dimensions[i+1].height = 18.75

# Merging Cell
for tag in range(2):
    ws[column[tag][0]+'2'] = '日付'
    ws.merge_cells(column[tag][0]+'2:'+column[tag][0]+'4') #日付
    ws[column[tag][1]+'2'] = '作業者印\n(始業)'
    ws.merge_cells(column[tag][1]+'2:'+column[tag][1]+'4') #作業者印
    ws[column[tag][2]+'2'] = '勤　　　　　務　　　　　内　　　　　容'
    ws.merge_cells(column[tag][2]+'2:'+column[tag][10]+'2') #勤務内容
    ws[column[tag][2]+'3'] = '作業内容'
    ws.merge_cells(column[tag][2]+'3:'+column[tag][2]+'4') #業内
    ws[column[tag][3]+'3'] = '定められた勤務時間'
    ws.merge_cells(column[tag][3]+'3:'+column[tag][3]+'4') #定められた勤務時間
    ws[column[tag][4]+'3'] = '減額時間'
    ws.merge_cells(column[tag][4]+'3:'+column[tag][4]+'4') #減額時間
    ws[column[tag][5]+'3'] = '(休憩時間)'
    ws.merge_cells(column[tag][5]+'3:'+column[tag][6]+'3') #休憩時間
    ws[column[tag][5]+'4'] = '超過勤務時間'
    ws.merge_cells(column[tag][5]+'4:'+column[tag][6]+'4') #超過勤務時間
    ws[column[tag][7]+'3'] = '超過勤務等'
    ws[column[tag][7]+'4'] = '100/100'
    ws[column[tag][8]+'4'] = '125/100'
    ws[column[tag][9]+'4'] = '25/100'
    ws.merge_cells(column[tag][7]+'3:'+column[tag][9]+'3') #超過勤務等
    ws[column[tag][10]+'3'] = 'その他'
    ws.merge_cells(column[tag][10]+'3:'+column[tag][10]+'4') #その他
    ws[column[tag][11]+'2'] = '監督・命令者\n認印'
    ws.merge_cells(column[tag][11]+'2:'+column[tag][11]+'4') #監督・命令者認印
    ws[column[tag][12]+'2'] = '従事者印\n(超勤)'
    ws.merge_cells(column[tag][12]+'2:'+column[tag][12]+'4') #従事者印

    if tag == 0:
        for i in range(13):
            for j in range(16):
                if i == 5:
                    ws[column[tag][i]+str(5+2*j)] = '(　：　)'
                    ws[column[tag][i]+str(6+2*j)] = '：'
                    #r_1 = column[tag][i] + str(5+2*j) + ':' + column[tag][i+1] + str(5+2*j)
                    #r_2 = column[tag][i] + str(6+2*j) + ':' + column[tag][i+1] + str(6+2*j)
                    #ws.merge_cells(r_1)
                    #ws.merge_cells(r_2)
                elif i == 6:
                    ws[column[tag][i]+str(5+2*j)] = '(　：　)'
                    ws[column[tag][i]+str(6+2*j)] = '：'
                else:
                    ws[column[tag][0]+str(5+2*j)] = str(j+1)
                    r = column[tag][i] + str(5+2*j) + ':' + column[tag][i] + str(6+2*j)
                    ws.merge_cells(r)
            for j in range(16):
                if isBizDay(int(year), int(month), j+1) == 0:
                    ws[column[tag][2]+str(5+2*j)] = '(祝 日)'
                if calendar.weekday(int(year), int(month), j+1) == 5:
                    ws[column[tag][2]+str(5+2*j)] = '(土 曜)'
                if calendar.weekday(int(year), int(month), j+1) == 6:
                    ws[column[tag][2]+str(5+2*j)] = '(日 曜)'
                ws[column[tag][0]+str(5+2*j)] = str(j+1)

    if tag == 1:
        for i in range(13):
            for j in range(15):
                if i == 5:
                    ws[column[tag][i]+str(5+2*j)] = '(　：　)'
                    ws[column[tag][i]+str(6+2*j)] = '：'
                    #r_1 = column[tag][i] + str(5+2*j) + ':' + column[tag][i+1] + str(5+2*j)
                    #r_2 = column[tag][i] + str(6+2*j) + ':' + column[tag][i+1] + str(6+2*j)
                    #ws.merge_cells(r_1)
                    #ws.merge_cells(r_2)
                elif i == 6:
                    ws[column[tag][i]+str(5+2*j)] = '(　：　)'
                    ws[column[tag][i]+str(6+2*j)] = '：'
                else:
                    r = column[tag][i] + str(5+2*j) + ':' +column[tag][i] + str(6+2*j)
                    ws.merge_cells(r)
            for j in range(calendar.monthrange(int(year), int(month))[1]-16):
                if isBizDay(int(year), int(month), j+17) == 0:
                    ws[column[tag][2]+str(5+2*j)] = '(祝 日)'
                if calendar.weekday(int(year), int(month), j+17) == 5:
                    ws[column[tag][2]+str(5+2*j)] = '(土 曜)'
                if calendar.weekday(int(year), int(month), j+17) == 6:
                    ws[column[tag][2]+str(5+2*j)] = '(日 曜)'
                ws[column[tag][0]+str(5+2*j)] = str(j+17)
        #最後の列
        ws[column[tag][0]+'35'] = '計'
        ws.merge_cells(column[tag][0]+'35:'+column[tag][2]+'36')
        ws.merge_cells(column[tag][3]+'35:'+column[tag][3]+'36')
        ws.merge_cells(column[tag][4]+'35:'+column[tag][4]+'36')
        ws.merge_cells(column[tag][5]+'35:'+column[tag][6]+'36')
        ws.merge_cells(column[tag][7]+'35:'+column[tag][7]+'36')
        ws.merge_cells(column[tag][8]+'35:'+column[tag][8]+'36')
        ws.merge_cells(column[tag][9]+'35:'+column[tag][9]+'36')
        ws.merge_cells(column[tag][10]+'35:'+column[tag][10]+'36')
        ws.merge_cells(column[tag][11]+'35:'+column[tag][12]+'36')



thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for row in ws:
    for cell in row:
        cell.alignment = Alignment(horizontal = 'center', vertical ='center', wrap_text=True)
        cell.border = thin_border

wb.save("test"+str(year)+str(month)+".xlsx")